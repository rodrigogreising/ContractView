import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import json
import uuid

from openpyxl import load_workbook

from .artifacts import Artifact, read_and_verify_artifact
from .ingestion import IngestionJob
from .provenance import LineageInput, append_lineage_tx

from ..ports.statements import Statement
from ..transaction import transaction as database
IMPORTER_VERSION = "ledger-importer-v1"
SCHEMA_VERSION = "expense-row-v1"
MAPPING_VERSION = "poc-ledger-mapping-v1"
HEADERS = ["expense_id","expense_date","vendor","description","budget_category","amount","invoice_number","evidence_file"]


class LedgerImportError(ValueError):
    pass


def _normalize_row(values, row_number: int) -> dict:
    row = dict(zip(HEADERS, values))
    if any(row.get(name) in (None, "") for name in HEADERS):
        raise LedgerImportError(f"Row {row_number} has missing required values")
    try:
        amount = Decimal(str(row["amount"])).quantize(Decimal("0.01"))
    except InvalidOperation as error:
        raise LedgerImportError(f"Row {row_number} amount is invalid") from error
    raw_date = row["expense_date"]
    if isinstance(raw_date, datetime): raw_date = raw_date.date()
    if isinstance(raw_date, date): normalized_date = raw_date.isoformat()
    else:
        try: normalized_date = date.fromisoformat(str(raw_date)).isoformat()
        except ValueError as error: raise LedgerImportError(f"Row {row_number} date is invalid") from error
    row["expense_date"] = normalized_date
    row["amount"] = amount
    return row


def parse_ledger(filename: str, content: bytes) -> tuple[str, list[tuple[int, dict, dict]]]:
    if filename.lower().endswith(".csv"):
        try: reader = csv.reader(StringIO(content.decode("utf-8-sig")))
        except UnicodeDecodeError as error: raise LedgerImportError("CSV must be UTF-8") from error
        rows = list(reader)
        sheet = "CSV"
        header_index = 0
    elif filename.lower().endswith(".xlsx"):
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active
            sheet = worksheet.title
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            workbook.close()
            normalized_rows = [[str(value).strip().lower().replace(" ", "_") for value in row] for row in rows]
            header_index = next((index for index, row in enumerate(normalized_rows) if row == HEADERS), -1)
        except Exception as error:
            raise LedgerImportError("XLSX workbook is malformed") from error
    else:
        raise LedgerImportError("Only CSV and XLSX ledgers are supported")
    if not rows or header_index < 0 or [str(value).strip().lower().replace(" ", "_") for value in rows[header_index]] != HEADERS:
        raise LedgerImportError("Ledger headers do not match expense-row-v1")
    parsed = []
    for row_number, values in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(value not in (None, "") for value in values): continue
        if str(values[0]).strip().lower() == "ledger control total": break
        if len(values) != len(HEADERS): raise LedgerImportError(f"Row {row_number} has the wrong column count")
        cells = {name: f"{chr(65 + index)}{row_number}" for index, name in enumerate(HEADERS)}
        parsed.append((row_number, _normalize_row(values, row_number), cells))
    if not parsed: raise LedgerImportError("Ledger contains no expense rows")
    keys = [row[1]["expense_id"] for row in parsed]
    if len(keys) != len(set(keys)): raise LedgerImportError("Ledger contains duplicate expense identifiers")
    return sheet, parsed


def import_ledger(job: IngestionJob, artifact: Artifact) -> dict:
    content = read_and_verify_artifact(artifact)
    sheet, rows = parse_ledger(artifact.filename, content)
    imported_total = sum((row[1]["amount"] for row in rows), Decimal("0.00"))
    with database() as connection:
        existing = connection.extraction.execute(Statement.LEDGER_IMPORT_READ_LEDGER_IMPORTS_001, (artifact.id,)
        ).fetchone()
        if existing:
            return {"id": existing[0], "rowCount": existing[1], "controlTotal": str(existing[2]), "importedTotal": str(existing[3])}
        config = connection.configuration.execute(Statement.LEDGER_IMPORT_READ_CONFIGURATION_VERSIONS_002, (job.contract_id,)
        ).fetchone()
        if not config: raise LedgerImportError("Activate reimbursement configuration before importing a ledger")
        try: control_total = Decimal(str(config[0]["ledgerControlTotal"])).quantize(Decimal("0.01"))
        except (KeyError, InvalidOperation) as error: raise LedgerImportError("Active configuration lacks a valid ledger control total") from error
        if imported_total != control_total:
            raise LedgerImportError(f"Ledger total {imported_total} does not reconcile to control total {control_total}")
        import_id = f"ledger-import-{uuid.uuid4().hex}"
        connection.extraction.execute(Statement.LEDGER_IMPORT_WRITE_LEDGER_IMPORTS_003,
            (import_id,job.id,artifact.id,job.contract_id,job.organization_id,sheet,IMPORTER_VERSION,SCHEMA_VERSION,MAPPING_VERSION,control_total,imported_total,len(rows)),
        )
        for row_number, row, cells in rows:
            expense_id = f"{import_id}:{row['expense_id']}"
            connection.extraction.execute(Statement.LEDGER_IMPORT_WRITE_EXPENSE_ROWS_004,
                (expense_id,import_id,job.contract_id,job.organization_id,row["expense_id"],row["expense_date"],row["vendor"],row["description"],
                 row["budget_category"],row["amount"],row["invoice_number"],row["evidence_file"],artifact.id,sheet,row_number,json.dumps(cells),
                 IMPORTER_VERSION,SCHEMA_VERSION,MAPPING_VERSION),
            )
            for field in ("expense_date","vendor","budget_category","amount","invoice_number","evidence_file"):
                append_lineage_tx(connection, LineageInput(
                    job.contract_id, job.organization_id, f"{row['expense_id']}.{field}",
                    str(row[field]), artifact.id, f"{sheet}!{cells[field]}", IMPORTER_VERSION,
                    mapping_version=MAPPING_VERSION,
                ))
        connection.commit()
    return {"id": import_id, "rowCount": len(rows), "controlTotal": str(control_total), "importedTotal": str(imported_total)}

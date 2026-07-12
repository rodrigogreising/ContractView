"""OpenPyXL spreadsheet parsing adapter."""

from io import BytesIO

from openpyxl import load_workbook

from ...application.ports.ledger_documents import SpreadsheetRows


class OpenpyxlSpreadsheetParser:
    def parse_xlsx(self, content: bytes) -> SpreadsheetRows:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            return SpreadsheetRows(
                sheet=worksheet.title,
                rows=[list(row) for row in worksheet.iter_rows(values_only=True)],
            )
        finally:
            workbook.close()
